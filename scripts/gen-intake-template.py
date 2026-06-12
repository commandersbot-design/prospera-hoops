"""Generate a coach-facing intake workbook for a team.

Pre-fills the roster from public/data/capitolHoops.json and lays out Schedule +
Box Score sheets with the EXACT column tokens scripts/ingest.mjs expects, so a
returned file converts straight to CSV and ingests cleanly.

    python scripts/gen-intake-template.py --team hayfield --out docs/Hayfield-Intake-Template.xlsx
"""
import argparse, json, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

ORANGE = "FF6A1A"; GRAPHITE = "0B0E13"; PANEL = "14181E"; YELLOW = "FFF3D6"; LIGHT = "F4F5F6"
FONT = "Arial"
hdr_font  = Font(name=FONT, bold=True, color="FFFFFF", size=11)
body_font = Font(name=FONT, size=11)
hdr_fill  = PatternFill("solid", fgColor=ORANGE)
fill_pre  = PatternFill("solid", fgColor=LIGHT)
fill_todo = PatternFill("solid", fgColor=YELLOW)
thin = Side(style="thin", color="D5D8DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

ap = argparse.ArgumentParser()
ap.add_argument("--team", default="hayfield")
ap.add_argument("--out", default=None)
a = ap.parse_args()

ch = json.load(open("public/data/capitolHoops.json", encoding="utf8"))
team = next((t for t in ch["teams"].values() if re.search(a.team, t["name"], re.I) or re.search(a.team, t.get("slug",""), re.I)), None)
if not team: raise SystemExit(f'No team matching "{a.team}"')
players = team["players"]
out = a.out or f"docs/{re.sub(r'[^A-Za-z0-9]+','-',team['name']).strip('-')}-Intake-Template.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)

wb = Workbook()

def style_header(ws, headers, comments=None, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
        if comments and h in comments:
            cm = Comment(comments[h], "Prospera Hoops"); cm.width = 220; cm.height = 90
            cell.comment = cm
    ws.freeze_panes = ws.cell(row=row+1, column=1)

# ---- Start Here ----
ws = wb.active; ws.title = "Start Here"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 104
def line(r, text, *, size=11, bold=False, color="11151C"):
    c = ws.cell(row=r, column=2, value=text)
    c.font = Font(name=FONT, size=size, bold=bold, color=color); c.alignment = left
    return r+1
r = 2
r = line(r, "PROSPERA HOOPS", size=20, bold=True, color=ORANGE)
r = line(r, f"{team['name']} — Player & Stats Intake", size=14, bold=True)
r += 1
r = line(r, "Thanks for helping us cover your team. Two things to fill in:", bold=True)
r = line(r, "1)  ROSTER tab — add each player's Height and Weight (yellow cells). Fix anything else that's wrong.")
r = line(r, "2)  For every game:  add one row on the SCHEDULE tab, and one row per player who played on the BOX SCORE tab.")
r += 1
r = line(r, "That's it — send the file back and we generate the player cards, recaps, leaderboard, and team page.", bold=True)
r += 1
r = line(r, "A FEW RULES (so nothing gets lost):", bold=True, color=ORANGE)
r = line(r, "•  Spell each player's name the same way every time — exactly as it appears on the Roster tab.")
r = line(r, "•  game_id is any short label you choose (G1, G2, …). Use the SAME id on the Schedule row and that game's Box Score rows — that's how they link.")
r = line(r, "•  DON'T enter points or rebounds — we calculate them (points = 2·made FGs + 3s made + made FTs; rebounds = off + def).")
r = line(r, "•  Leave a cell blank or 0 if it didn't happen. Minutes are optional but nice to have.")
r += 1
r = line(r, "BOX SCORE COLUMN KEY:", bold=True, color=ORANGE)
key = [
 ("min", "Minutes played"),("fgm / fga", "Field goals made / attempted (include 3-pointers)"),
 ("tpm / tpa", "3-pointers made / attempted"),("ftm / fta", "Free throws made / attempted"),
 ("oreb / dreb", "Offensive / defensive rebounds"),("ast", "Assists"),("stl", "Steals"),
 ("blk", "Blocks"),("tov", "Turnovers"),("pf", "Personal fouls"),
]
for k, v in key:
    c = ws.cell(row=r, column=2, value=f"     {k}  —  {v}"); c.font = Font(name=FONT, size=10.5); c.alignment = left; r += 1
r += 1
r = line(r, "CONSENT (please confirm before sharing player info):", bold=True, color=ORANGE)
r = line(r, "By sharing player information with Prospera Hoops, the program confirms it is authorized to do so and that "
            "players' parents or guardians consent to their athlete's name, stats, profile, and any provided photos or film "
            "appearing on Prospera Hoops and partner channels. A player or guardian may request edits or removal at any time "
            "by emailing info@prosperahoops.com, and we'll honor it promptly.", size=10)
r += 1
r = line(r, "Questions? info@prosperahoops.com", bold=True)

# ---- Roster ----
ws = wb.create_sheet("Roster")
ros_headers = ["team","number","player","pos","grad_year","height","weight","offers","commit","film_link","instagram"]
ros_comments = {
 "height":"Fill this in — e.g. 6'2\"","weight":"Fill this in — e.g. 185","pos":"PG, SG, G, W, F, C…",
 "grad_year":"HS graduation year, e.g. 2027","offers":"(optional) college offers, comma-separated",
 "commit":"(optional) committed school, if any","film_link":"(optional) Hudl / YouTube link",
 "instagram":"(optional) player or recruiting IG handle",
}
style_header(ws, ros_headers, ros_comments)
for i, p in enumerate(players, start=2):
    vals = [team["name"], p.get("number",""), p.get("name",""), p.get("position",""), p.get("classYear",""), "", "", "", "", "", ""]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=c, value=v); cell.font = body_font; cell.border = border
        cell.alignment = center if c in (1,2,4,5) else left
        if c <= 5: cell.fill = fill_pre
        if c in (6,7): cell.fill = fill_todo   # height/weight = fill me
widths = [16,8,22,7,10,10,9,26,16,28,18]
for c, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(c)].width = w

# ---- Schedule ----
ws = wb.create_sheet("Schedule")
sch_headers = ["game_id","date","event","opponent","home_away","team_pts","opp_pts"]
sch_comments = {
 "game_id":"Any short label (G1, G2…). Must match the Box Score rows for this game.",
 "date":"e.g. June 14 2026","event":"League / tournament name","home_away":"home or away",
 "team_pts":f"{team['name']} final score","opp_pts":"Opponent final score",
}
style_header(ws, sch_headers, sch_comments)
# G7 is set up as the coach's first game (the Box Score tab is pre-seeded with the
# same id). They fill date/opponent/score here and the numbers there.
first = ["G7","", "Capitol Hoops", "", "", "", ""]
for c, v in enumerate(first, 1):
    cell = ws.cell(row=2, column=c, value=v); cell.font = body_font; cell.border = border; cell.alignment = center
    if c in (1,3): cell.fill = fill_pre
    elif c in (2,4,5,6,7): cell.fill = fill_todo
ws.cell(row=4, column=1, value="↑ Your first game (G7). Add G8, G9… below as you play them — same id goes on the Box Score tab.").font = Font(name=FONT, italic=True, color=ORANGE)
for c, w in enumerate([10,16,18,24,12,11,10], 1): ws.column_dimensions[get_column_letter(c)].width = w

# ---- Box Score ----
ws = wb.create_sheet("Box Score")
box_headers = ["game_id","date","player","min","fgm","fga","tpm","tpa","ftm","fta","oreb","dreb","ast","stl","blk","tov","pf","started","dfl","chg"]
box_comments = {
 "game_id":"Same id you used on the Schedule tab for this game.","player":"Exactly as on the Roster tab.",
 "min":"Minutes played. PLEASE fill this in — it unlocks per-36 stats and minutes load.",
 "fgm":"Field goals made (include 3s)","fga":"Field goals attempted","tpm":"3-pointers made","tpa":"3-pointers attempted",
 "ftm":"Free throws made","fta":"Free throws attempted","oreb":"Offensive rebounds","dreb":"Defensive rebounds","tov":"Turnovers","pf":"Personal fouls",
 "started":"OPTIONAL: 1 if the player started, else 0.","dfl":"OPTIONAL: deflections.","chg":"OPTIONAL: charges drawn.",
}
style_header(ws, box_headers, box_comments)
# Pre-seed one game block with all rostered players so the coach just fills numbers.
opt_cols = set(range(len(box_headers)-2, len(box_headers)+1))  # started, dfl, chg
for i, p in enumerate(players, start=2):
    ws.cell(row=i, column=1, value="G7").font = body_font
    ws.cell(row=i, column=3, value=p.get("name","")).font = body_font
    for c in range(1, len(box_headers)+1):
        ws.cell(row=i, column=c).border = border
        ws.cell(row=i, column=c).alignment = center
        if c in (1,3): ws.cell(row=i, column=c).fill = fill_pre
        elif c == 4: ws.cell(row=i, column=c).fill = fill_todo   # minutes — please fill
        elif c in opt_cols: ws.cell(row=i, column=c).fill = fill_pre  # optional extras
note = ws.cell(row=len(players)+3, column=1,
    value="↑ One 15-row block = one game. For the next game, copy these name rows, change game_id (G8…) and date. Delete players who didn't play.")
note.font = Font(name=FONT, italic=True, color=ORANGE)
ws.column_dimensions["A"].width = 9; ws.column_dimensions["B"].width = 15; ws.column_dimensions["C"].width = 22
for c in range(4, len(box_headers)+1): ws.column_dimensions[get_column_letter(c)].width = 6.5

wb.save(out)
print("saved:", out, "| players:", len(players))
